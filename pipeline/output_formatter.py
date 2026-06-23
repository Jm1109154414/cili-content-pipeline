"""Genera el resumen final del mes: Excel, Markdown y checklist de publicación."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .brief_schema import ClientBrief
from .historial import guardar_contenido_final

COLUMNAS = [
    "No.", "Fecha", "Plataforma", "Tipo", "Objetivo", "Gancho",
    "Copy completo", "Hashtags", "CTA", "Ruta imagen", "Estado",
]


def _filas(contenido: dict) -> list[dict]:
    """contenido viene de chequeo-marca: ya consolidó por post_id el estado
    final y la ruta_imagen de estrategia-copy + imagenes (ver vocabulario
    único de estados e identificador de post en PLAN_MAESTRO.md).
    {"posts": [{post_id, fecha, plataforma, tipo, objetivo, gancho,
    copy_completo, hashtags, cta, ruta_imagen, estado, ...}]}."""
    posts_contenido = contenido.get("posts", [])

    filas = []
    for i, post in enumerate(posts_contenido):
        filas.append(
            {
                "No.": i + 1,
                "Fecha": post.get("fecha", ""),
                "Plataforma": post.get("plataforma", ""),
                "Tipo": post.get("tipo", ""),
                "Objetivo": post.get("objetivo", ""),
                "Gancho": post.get("gancho", ""),
                "Copy completo": post.get("copy_completo", ""),
                "Hashtags": ", ".join(post.get("hashtags", [])),
                "CTA": post.get("cta", ""),
                "Ruta imagen": post.get("ruta_imagen", ""),
                "Estado": post.get("estado", "GENERADO"),
            }
        )
    return filas


def _generar_xlsx(filas: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendario"
    ws.append(COLUMNAS)
    for fila in filas:
        ws.append([fila[c] for c in COLUMNAS])
    for idx, columna in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(15, len(columna) + 5)
    wb.save(path)


def _generar_md(filas: list[dict], brief: ClientBrief, path: Path) -> None:
    lineas = [f"# Calendario de contenido — {brief.empresa} — {brief.mes}", ""]
    for fila in filas:
        lineas.append(f"## Post {fila['No.']} — {fila['Plataforma']} ({fila['Fecha']})")
        lineas.append(f"- **Tipo:** {fila['Tipo']}")
        lineas.append(f"- **Objetivo:** {fila['Objetivo']}")
        lineas.append(f"- **Gancho:** {fila['Gancho']}")
        lineas.append(f"- **Copy completo:**\n\n{fila['Copy completo']}\n")
        lineas.append(f"- **Hashtags:** {fila['Hashtags']}")
        lineas.append(f"- **CTA:** {fila['CTA']}")
        lineas.append(f"- **Imagen:** {fila['Ruta imagen']}")
        lineas.append(f"- **Estado:** {fila['Estado']}")
        lineas.append("")
    path.write_text("\n".join(lineas), encoding="utf-8")


def _generar_checklist(filas: list[dict], path: Path) -> None:
    lineas = ["# Checklist de publicación", ""]
    for fila in filas:
        lineas.append(f"- [ ] Post {fila['No.']} — {fila['Plataforma']} ({fila['Fecha']}) — {fila['Estado']}")
    path.write_text("\n".join(lineas), encoding="utf-8")


def generate_output(contenido: dict, brief: ClientBrief, output_dir: str | Path) -> dict:
    """Genera resumen_calendario.xlsx, resumen_calendario.md, checklist_publicacion.md
    y contenido_final.json (historial para que meses futuros no repitan temas).

    contenido: salida final de chequeo-marca (estrategia+copy+imagen+estado
    ya consolidados por post_id, ver PLAN_MAESTRO.md).
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    filas = _filas(contenido)

    xlsx_path = output_dir / "resumen_calendario.xlsx"
    md_path = output_dir / "resumen_calendario.md"
    checklist_path = output_dir / "checklist_publicacion.md"

    _generar_xlsx(filas, xlsx_path)
    _generar_md(filas, brief, md_path)
    _generar_checklist(filas, checklist_path)
    historial_path = guardar_contenido_final(contenido, output_dir)

    return {
        "xlsx": str(xlsx_path),
        "md": str(md_path),
        "checklist": str(checklist_path),
        "historial": str(historial_path),
    }
